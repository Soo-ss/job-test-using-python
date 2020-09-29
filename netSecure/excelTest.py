import openpyxl

wb = openpyxl.load_workbook("test.xlsx")
ws = wb["Sheet1"]

# print(ws.cell(1, 2).value)

for row in ws.rows:
    for cell in row:
        print(str(cell.value) + ",", end="")
    print()

write_wb = openpyxl.Workbook()
write_ws = write_wb.active
write_ws.append([1, 2, 3])
write_ws.append([1, 2, 3])
write_ws.append([1, 2, 3])
write_wb.save("test1.xlsx")